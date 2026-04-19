from decimal import Decimal
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from database import get_db
from models import Pedido, PedidoCliente

router = APIRouter(prefix="/pedidos", tags=["export"])

GREEN = "FF92D050"
YELLOW = "FFFFFF00"
HEADER_BG = "FFD9E1F2"
TOTAL_BG = "FFFCE4D6"
PAGADO_BG = "FF92D050"
PENDIENTE_BG = "FFFFC000"

thin = Side(style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _estilo_celda(ws, row, col, value, bold=False, bg=None, align="left", number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold)
    cell.border = border
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if bg:
        cell.fill = PatternFill(fill_type="solid", fgColor=bg)
    if number_format:
        cell.number_format = number_format
    return cell


def _calcular(pc: PedidoCliente):
    comision_unit = Decimal(str(pc.comision_por_item if pc.comision_por_item is not None else pc.cliente.comision_por_item))
    items_activos = [i for i in pc.items if i.deleted_at is None]
    items_llegados = [i for i in items_activos if i.llegado]
    pagos_activos = [p for p in pc.pagos if p.deleted_at is None]
    subtotal = sum(Decimal(str(i.precio or 0)) for i in items_llegados)
    comision = Decimal(len(items_llegados)) * comision_unit
    total = subtotal + comision
    pagado = sum(Decimal(str(p.monto)) for p in pagos_activos)
    saldo = total - pagado
    return items_activos, items_llegados, pagos_activos, subtotal, comision, total, pagado, saldo


def _hoja_total(wb: Workbook, pedido: Pedido, pcs):
    ws = wb.active
    ws.title = "Total"

    total_subtotal = Decimal("0")
    total_comision = Decimal("0")
    total_pagado = Decimal("0")
    resumen_clientes = []

    for pc in pcs:
        items_activos, items_llegados, pagos_activos, subtotal, comision, total, pagado, saldo = _calcular(pc)
        total_subtotal += subtotal
        total_comision += comision
        total_pagado += pagado
        resumen_clientes.append({
            "nombre": pc.cliente.nombre,
            "items": len(items_llegados),
            "subtotal": subtotal,
            "comision": comision,
            "total": total,
            "pagado": pagado,
            "saldo": saldo,
            "estado": "PAGADO" if saldo <= 0 else "PENDIENTE",
        })

    # Fila 1: info del pedido
    ws.merge_cells("A1:B1")
    ws["A1"] = f"Pedido #{pedido.numero}" if pedido.numero else "Pedido"
    ws["A1"].font = Font(bold=True, size=12)
    ws["C1"] = str(pedido.fecha)
    ws["C1"].font = Font(bold=True)
    ws["E1"] = float(total_subtotal)
    ws["E1"].number_format = '#,##0.00'
    ws["F1"] = "Subtotal total"
    ws["G1"] = float(total_comision)
    ws["G1"].number_format = '#,##0.00'
    ws["H1"] = "Comision total"

    # Fila 2: encabezados
    headers = ["Num.", "Cliente", "Items llegados", "Subtotal", "Comision", "Total", "Pagado", "Saldo", "Estado"]
    for col, h in enumerate(headers, 1):
        _estilo_celda(ws, 2, col, h, bold=True, bg=HEADER_BG, align="center")

    # Filas de clientes
    for i, c in enumerate(resumen_clientes, 1):
        row = i + 2
        bg_estado = PAGADO_BG if c["estado"] == "PAGADO" else PENDIENTE_BG
        _estilo_celda(ws, row, 1, i, align="center")
        _estilo_celda(ws, row, 2, c["nombre"], bold=True)
        _estilo_celda(ws, row, 3, c["items"], align="center")
        _estilo_celda(ws, row, 4, float(c["subtotal"]), number_format='#,##0.00')
        _estilo_celda(ws, row, 5, float(c["comision"]), number_format='#,##0.00')
        _estilo_celda(ws, row, 6, float(c["total"]), number_format='#,##0.00')
        _estilo_celda(ws, row, 7, float(c["pagado"]), number_format='#,##0.00')
        _estilo_celda(ws, row, 8, float(c["saldo"]), number_format='#,##0.00')
        _estilo_celda(ws, row, 9, c["estado"], bold=True, bg=bg_estado, align="center")

    # Fila de totales
    summary_row = len(resumen_clientes) + 3
    ws.cell(row=summary_row, column=3, value="TOTALES").font = Font(bold=True)
    _estilo_celda(ws, summary_row, 4, float(total_subtotal), bold=True, bg=TOTAL_BG, number_format='#,##0.00')
    _estilo_celda(ws, summary_row, 5, float(total_comision), bold=True, bg=TOTAL_BG, number_format='#,##0.00')
    total_general = total_subtotal + total_comision
    _estilo_celda(ws, summary_row, 6, float(total_general), bold=True, bg=TOTAL_BG, number_format='#,##0.00')
    _estilo_celda(ws, summary_row, 7, float(total_pagado), bold=True, bg=TOTAL_BG, number_format='#,##0.00')
    saldo_total = total_general - total_pagado
    bg_saldo = PAGADO_BG if saldo_total <= 0 else PENDIENTE_BG
    _estilo_celda(ws, summary_row, 8, float(saldo_total), bold=True, bg=bg_saldo, number_format='#,##0.00')

    # Anchos de columna
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 12


def _hoja_cliente(wb: Workbook, pc: PedidoCliente):
    nombre = pc.cliente.nombre[:31]  # Excel limita a 31 chars el nombre de hoja
    ws = wb.create_sheet(title=nombre)

    items_activos, items_llegados, pagos_activos, subtotal, comision, total, pagado, saldo = _calcular(pc)

    # Fila 1: info del cliente
    _estilo_celda(ws, 1, 1, pc.cliente.nombre, bold=True, bg=HEADER_BG)
    _estilo_celda(ws, 1, 2, float(subtotal), bold=True, number_format='#,##0.00')
    _estilo_celda(ws, 1, 3, float(pc.comision_por_item or 0), number_format='#,##0.00')
    ws.cell(row=1, column=4, value="$/item comision")

    # Fila 2: encabezados items
    headers = ["Num.", "Link", "Articulo", "Llegado", "Precio"]
    for col, h in enumerate(headers, 1):
        _estilo_celda(ws, 2, col, h, bold=True, bg=HEADER_BG, align="center")

    # Filas de items
    for i, item in enumerate(items_activos, 1):
        row = i + 2
        llegado = "V" if item.llegado else "X"
        bg_llegado = PAGADO_BG if item.llegado else None
        _estilo_celda(ws, row, 1, item.numero or i, align="center")
        _estilo_celda(ws, row, 2, item.link or "")
        _estilo_celda(ws, row, 3, item.articulo or "")
        _estilo_celda(ws, row, 4, llegado, bg=bg_llegado, align="center")
        _estilo_celda(ws, row, 5, float(item.precio or 0) if item.llegado else "", number_format='#,##0.00')

    # Filas de resumen
    base = len(items_activos) + 3
    _estilo_celda(ws, base, 4, float(subtotal), number_format='#,##0.00')
    _estilo_celda(ws, base, 5, "Sub total", bold=True, bg=HEADER_BG)
    _estilo_celda(ws, base + 1, 4, float(comision), number_format='#,##0.00')
    _estilo_celda(ws, base + 1, 5, "Comision", bold=True, bg=HEADER_BG)
    _estilo_celda(ws, base + 2, 4, float(total), bold=True, number_format='#,##0.00')
    _estilo_celda(ws, base + 2, 5, "Total", bold=True, bg=TOTAL_BG)

    # Pagos
    if pagos_activos:
        pago_row = base + 4
        ws.cell(row=pago_row, column=1, value="Pagos").font = Font(bold=True)
        headers_pagos = ["Tipo", "Monto", "Fecha", "Notas"]
        for col, h in enumerate(headers_pagos, 1):
            _estilo_celda(ws, pago_row + 1, col, h, bold=True, bg=HEADER_BG, align="center")
        for j, pago in enumerate(pagos_activos):
            r = pago_row + 2 + j
            _estilo_celda(ws, r, 1, pago.tipo or "")
            _estilo_celda(ws, r, 2, float(pago.monto), number_format='#,##0.00')
            _estilo_celda(ws, r, 3, str(pago.fecha.date()) if pago.fecha else "")
            _estilo_celda(ws, r, 4, pago.notas or "")
        saldo_row = pago_row + 2 + len(pagos_activos) + 1
    else:
        saldo_row = base + 5

    bg_saldo = PAGADO_BG if saldo <= 0 else PENDIENTE_BG
    _estilo_celda(ws, saldo_row, 4, float(saldo), bold=True, number_format='#,##0.00')
    _estilo_celda(ws, saldo_row, 5, "Saldo", bold=True, bg=bg_saldo)

    # Anchos de columna
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14


@router.get("/{pedido_id}/export")
def exportar_pedido_excel(pedido_id: int, db: Session = Depends(get_db)):
    pedido = (
        db.query(Pedido)
        .options(
            joinedload(Pedido.pedido_clientes).joinedload(PedidoCliente.cliente),
            joinedload(Pedido.pedido_clientes).joinedload(PedidoCliente.items),
            joinedload(Pedido.pedido_clientes).joinedload(PedidoCliente.pagos),
        )
        .filter(Pedido.id == pedido_id, Pedido.deleted_at.is_(None))
        .first()
    )
    if not pedido:
        raise HTTPException(status_code=404, detail="Pedido no encontrado")

    pcs = [pc for pc in pedido.pedido_clientes if pc.deleted_at is None]

    wb = Workbook()
    _hoja_total(wb, pedido, pcs)
    for pc in pcs:
        _hoja_cliente(wb, pc)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    numero = pedido.numero or pedido_id
    filename = f"Pedido_{numero}_{pedido.fecha}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
